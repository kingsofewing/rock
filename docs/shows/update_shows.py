#!/usr/bin/env python3
# Run with: python3 docs/shows/update_shows.py
"""
update_shows.py
---------------
Reads shows.xlsx and writes shows.json (both in docs/shows/).
Run this whenever you update the spreadsheet:

    python3 docs/shows/update_shows.py

Then commit shows.xlsx and shows.json via GitHub Desktop and push.
Requires openpyxl:  pip3 install openpyxl
"""

import datetime
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("Missing dependency. Run:  pip3 install openpyxl")

THIS_DIR  = Path(__file__).resolve().parent   # docs/shows/
XLSX_PATH = THIS_DIR / "shows.xlsx"
JSON_PATH = THIS_DIR / "shows.json"

# Column positions (1-based) in the Shows sheet
COL_DATE             = 1
COL_VENUE_NAME       = 2
COL_VENUE_URL        = 3
COL_START_TIME       = 4
COL_END_TIME         = 5
COL_LOCATION         = 6
COL_ADDRESS          = 7
COL_DETAIL           = 8
COL_DETAIL_LINK_NAME = 9
COL_DETAIL_LINK_URL  = 10
COL_DETAIL_LINK_SUFFIX = 11
COL_DESCRIPTION      = 12
COL_CANCELLED        = 13
COL_CANCEL_REASON    = 14
COL_PRIVATE          = 15
COL_INSTAGRAM_LINK   = 16

DATA_START_ROW = 3  # row 1 = instructions, row 2 = headers


def cell_val(row, col):
    v = row[col - 1].value
    if v is None:
        return ""
    # Excel time cells come back as datetime.time — convert to "H:MM AM/PM"
    if isinstance(v, datetime.time):
        return v.strftime("%-I:%M %p")   # e.g. "4:00 PM"
    if isinstance(v, datetime.datetime):
        return v.strftime("%-I:%M %p")
    return str(v).strip()


def build_shows():
    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb["Shows"]
    shows = []

    for row in ws.iter_rows(min_row=DATA_START_ROW):
        date = cell_val(row, COL_DATE)
        if not date:
            continue  # skip blank rows

        show = {"date": date}

        venue_name = cell_val(row, COL_VENUE_NAME)
        venue_url  = cell_val(row, COL_VENUE_URL)
        if venue_name:
            show["venue_name"] = venue_name
        if venue_url:
            show["venue_url"] = venue_url

        start_time = cell_val(row, COL_START_TIME)
        end_time   = cell_val(row, COL_END_TIME)
        location   = cell_val(row, COL_LOCATION)
        if start_time:
            show["start_time"] = start_time
        if end_time:
            show["end_time"] = end_time
        if location:
            show["location"] = location

        address = cell_val(row, COL_ADDRESS)
        if address:
            show["address"] = address

        detail = cell_val(row, COL_DETAIL)
        if detail:
            show["detail"] = detail

        dl_name   = cell_val(row, COL_DETAIL_LINK_NAME)
        dl_url    = cell_val(row, COL_DETAIL_LINK_URL)
        dl_suffix = cell_val(row, COL_DETAIL_LINK_SUFFIX)
        if dl_name:
            show["detail_link_name"] = dl_name
        if dl_url:
            show["detail_link_url"] = dl_url
        if dl_suffix:
            show["detail_link_suffix"] = dl_suffix.lstrip()

        description = cell_val(row, COL_DESCRIPTION)
        if description:
            show["description"] = description

        cancelled = cell_val(row, COL_CANCELLED).upper()
        if cancelled == "TRUE":
            show["cancelled"] = True

        cancel_reason = cell_val(row, COL_CANCEL_REASON)
        if cancel_reason:
            show["cancel_reason"] = cancel_reason

        private = cell_val(row, COL_PRIVATE).upper()
        if private == "TRUE":
            show["private"] = True

        instagram_link = cell_val(row, COL_INSTAGRAM_LINK)
        if instagram_link:
            show["instagram_link"] = instagram_link

        shows.append(show)

    return shows


def main():
    if not XLSX_PATH.exists():
        sys.exit(f"Cannot find {XLSX_PATH}\nMake sure you're running from the repo root.")

    shows = build_shows()
    JSON_PATH.write_text(json.dumps(shows, indent=2, ensure_ascii=False))
    print(f"Wrote {len(shows)} shows to {JSON_PATH.name}")


if __name__ == "__main__":
    main()
