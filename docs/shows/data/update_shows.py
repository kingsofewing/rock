#!/usr/bin/env python3
# Run with: python3 docs/shows/data/update_shows.py
"""
update_shows.py
---------------
Reads shows.xlsx and writes shows.json (both in docs/shows/data/).
Run this whenever you update the spreadsheet:

    python3 docs/shows/data/update_shows.py

Then commit shows.xlsx and shows.json via GitHub Desktop and push.
Requires openpyxl:  pip3 install openpyxl
"""

import datetime
import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("Missing dependency. Run:  pip3 install openpyxl")

THIS_DIR   = Path(__file__).resolve().parent      # docs/shows/data/
SHOWS_DIR  = THIS_DIR.parent                       # docs/shows/
REPO_ROOT  = SHOWS_DIR.parent.parent               # repo root
XLSX_PATH  = SHOWS_DIR / "data" / "shows.xlsx"
JSON_PATH  = SHOWS_DIR / "data" / "shows.json"
ICS_DIR    = SHOWS_DIR / "calendar"               # hosted per-gig .ics files
PROMO_DIR  = SHOWS_DIR / "promos"                 # hosted per-gig promo images
SOCIAL_DIR = SHOWS_DIR.parent / "social media"    # source gig folders (YYYYMMDD Venue)
SITE_URL  = "https://www.kingsofewing.com"
IMG_EXTS  = (".png", ".jpg", ".jpeg", ".webp")

# Column positions (1-based) in the Shows sheet
COL_DATE             = 1
COL_VENUE_NAME       = 2
COL_VENUE_URL        = 3
COL_START_TIME       = 4
COL_END_TIME         = 5
COL_LOCATION         = 6
COL_ADDRESS          = 7
COL_DETAIL           = 8
COL_DETAIL_LINK_NAME = 9
COL_DETAIL_LINK_URL  = 10
COL_DETAIL_LINK_SUFFIX = 11
COL_DESCRIPTION      = 12
COL_CANCELLED        = 13
COL_CANCEL_REASON    = 14
COL_PRIVATE          = 15
COL_INSTAGRAM_LINK   = 16
# Columns 17-21 hold weather data. Ticket info appended after it.
COL_TICKET_PRICE     = 22
COL_TICKET_URL       = 23
COL_TICKET_NOTE      = 24

DATA_START_ROW = 3  # row 1 = instructions, row 2 = headers


def cell_val(row, col):
    v = row[col - 1].value
    if v is None:
        return ""
    # Excel time cells come back as datetime.time — convert to "H:MM AM/PM"
    if isinstance(v, datetime.time):
        return v.strftime("%-I:%M %p")   # e.g. "4:00 PM"
    if isinstance(v, datetime.datetime):
        return v.strftime("%-I:%M %p")
    return str(v).strip()


def build_shows():
    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb["Shows"]
    shows = []

    for row in ws.iter_rows(min_row=DATA_START_ROW):
        date = cell_val(row, COL_DATE)
        if not date:
            continue  # skip blank rows

        show = {"date": date}

        venue_name = cell_val(row, COL_VENUE_NAME)
        venue_url  = cell_val(row, COL_VENUE_URL)
        if venue_name:
            show["venue_name"] = venue_name
        if venue_url:
            show["venue_url"] = venue_url

        start_time = cell_val(row, COL_START_TIME)
        end_time   = cell_val(row, COL_END_TIME)
        location   = cell_val(row, COL_LOCATION)
        if start_time:
            show["start_time"] = start_time
        if end_time:
            show["end_time"] = end_time
        if location:
            show["location"] = location

        address = cell_val(row, COL_ADDRESS)
        if address:
            show["address"] = address

        detail = cell_val(row, COL_DETAIL)
        if detail:
            show["detail"] = detail

        dl_name   = cell_val(row, COL_DETAIL_LINK_NAME)
        dl_url    = cell_val(row, COL_DETAIL_LINK_URL)
        dl_suffix = cell_val(row, COL_DETAIL_LINK_SUFFIX)
        if dl_name:
            show["detail_link_name"] = dl_name
        if dl_url:
            show["detail_link_url"] = dl_url
        if dl_suffix:
            show["detail_link_suffix"] = dl_suffix.lstrip()

        description = cell_val(row, COL_DESCRIPTION)
        if description:
            show["description"] = description

        cancelled = cell_val(row, COL_CANCELLED).upper()
        if cancelled == "TRUE":
            show["cancelled"] = True

        cancel_reason = cell_val(row, COL_CANCEL_REASON)
        if cancel_reason:
            show["cancel_reason"] = cancel_reason

        private = cell_val(row, COL_PRIVATE).upper()
        if private == "TRUE":
            show["private"] = True

        instagram_link = cell_val(row, COL_INSTAGRAM_LINK)
        if instagram_link:
            show["instagram_link"] = instagram_link

        # Ticket price drives the JSON-LD "offers" block. Leave blank when
        # unknown — the site then omits offers entirely rather than telling
        # Google the show is free. "Free" or "0" means genuinely free.
        ticket_price = cell_val(row, COL_TICKET_PRICE)
        if ticket_price:
            cleaned = ticket_price.replace("$", "").strip()
            if cleaned.lower() in ("free", "0", "0.00"):
                show["ticket_price"] = "0"
            else:
                show["ticket_price"] = cleaned

        ticket_url = cell_val(row, COL_TICKET_URL)
        if ticket_url:
            show["ticket_url"] = ticket_url

        # Short qualifier shown in parentheses after the price, e.g.
        # "includes a free drink" -> "$10 (includes a free drink)".
        ticket_note = cell_val(row, COL_TICKET_NOTE)
        if ticket_note:
            show["ticket_note"] = ticket_note

        shows.append(show)

    return shows


# ─────────────────────────────────────────────────────────────
# Calendar (.ics) generation — universal format that Apple
# Calendar, Outlook, and Google all open. Mirrors the buildICS
# logic in index.html so the email links and the website button
# stay consistent. Files are hosted at:
#   {SITE_URL}/docs/shows/calendar/kings-of-ewing-YYYY-MM-DD.ics
# ─────────────────────────────────────────────────────────────

def _parse_time(s):
    """Parse 'H:MM AM/PM' -> (hour24, minute) or None."""
    if not s:
        return None
    m = re.search(r"(\d{1,2}):(\d{2})\s*(AM|PM)", s, re.I)
    if not m:
        return None
    h = int(m.group(1)) % 12
    if m.group(3).upper() == "PM":
        h += 12
    return h, int(m.group(2))


def _chicago_offset(date_str):
    """UTC offset (timedelta) for America/Chicago on the given date (handles CST/CDT)."""
    try:
        from zoneinfo import ZoneInfo
        d = datetime.datetime.fromisoformat(date_str + "T12:00:00").replace(
            tzinfo=ZoneInfo("America/Chicago"))
        return d.utcoffset()
    except Exception:
        return datetime.timedelta(hours=-6)


def _start_end(show):
    """Return (start_dt_utc, end_dt_utc, has_time). ~2hr default set length."""
    date_str = show["date"]
    t = _parse_time(show.get("start_time")) or _parse_time(show.get("detail"))
    if not t:
        return None, None, False
    off = _chicago_offset(date_str)
    y, mo, da = (int(x) for x in date_str.split("-"))
    start_local = datetime.datetime(y, mo, da, t[0], t[1])
    te = _parse_time(show.get("end_time"))
    if te:
        end_local = datetime.datetime(y, mo, da, te[0], te[1])
        if end_local <= start_local:
            end_local += datetime.timedelta(days=1)
    else:
        end_local = start_local + datetime.timedelta(hours=2)
    tz = datetime.timezone(off)
    return (start_local.replace(tzinfo=tz).astimezone(datetime.timezone.utc),
            end_local.replace(tzinfo=tz).astimezone(datetime.timezone.utc),
            True)


def _ics_esc(txt):
    return (str(txt).replace("\\", "\\\\").replace(";", "\\;")
            .replace(",", "\\,").replace("\r\n", "\\n").replace("\n", "\\n"))


def _fold(line):
    """RFC 5545 line folding at 75 octets."""
    if len(line) <= 74:
        return line
    out, rest = line[:74], line[74:]
    while len(rest) > 73:
        out += "\r\n " + rest[:73]
        rest = rest[73:]
    return out + "\r\n " + rest


def build_ics(show):
    venue = show.get("venue_name") or "TBA"
    start, end, has_time = _start_end(show)
    if has_time:
        fmt = lambda d: d.strftime("%Y%m%dT%H%M%SZ")
        dt_start = "DTSTART:" + fmt(start)
        dt_end   = "DTEND:" + fmt(end)
    else:
        d0 = datetime.date.fromisoformat(show["date"])
        d1 = d0 + datetime.timedelta(days=1)
        dt_start = "DTSTART;VALUE=DATE:" + d0.strftime("%Y%m%d")
        dt_end   = "DTEND;VALUE=DATE:" + d1.strftime("%Y%m%d")

    loc_parts = [venue]
    if show.get("location"):
        loc_parts.append(show["location"])
    if show.get("address"):
        loc_parts.append(show["address"])
    location = ", ".join(loc_parts)
    desc = "\n\n".join(filter(None, [show.get("description"), f"More info: {SITE_URL}"]))
    uid = f"{show['date']}-" + re.sub(r"[^a-z0-9]+", "-", venue.lower()) + "@kingsofewing.com"
    stamp = datetime.datetime.now(datetime.timezone.utc).strftime("%Y%m%dT%H%M%SZ")

    lines = [
        "BEGIN:VCALENDAR", "VERSION:2.0",
        "PRODID:-//Kings of Ewing//Website//EN",
        "CALSCALE:GREGORIAN", "METHOD:PUBLISH", "BEGIN:VEVENT",
        "UID:" + uid, "DTSTAMP:" + stamp, dt_start, dt_end,
        "SUMMARY:" + _ics_esc("Kings of Ewing at " + venue),
        "LOCATION:" + _ics_esc(location),
        "DESCRIPTION:" + _ics_esc(desc),
        "URL:" + (show.get("venue_url") or SITE_URL),
        "END:VEVENT", "END:VCALENDAR",
    ]
    return "\r\n".join(_fold(l) for l in lines) + "\r\n"


def write_ics_files(shows):
    """Generate one .ics per public (non-private) show into ICS_DIR."""
    ICS_DIR.mkdir(exist_ok=True)
    count = 0
    for s in shows:
        if s.get("private"):
            continue
        path = ICS_DIR / f"kings-of-ewing-{s['date']}.ics"
        path.write_text(build_ics(s), newline="")
        count += 1
    return count


# ─────────────────────────────────────────────────────────────
# Promo images — copy each gig's promo graphic from its social-media
# folder (docs/social media/YYYYMMDD Venue/) to a clean hosted path
# (docs/shows/promos/YYYY-MM-DD.<ext>) so gig emails can embed it inline
# at {SITE_URL}/docs/shows/promos/YYYY-MM-DD.<ext>. Promos are designed
# by hand and dropped in the gig folder; this just publishes them to a
# tidy, space-free URL keyed by date.
# ─────────────────────────────────────────────────────────────

def _find_promo(date_str):
    """Find the promo image in the social-media folder for this date, if any."""
    if not SOCIAL_DIR.is_dir():
        return None
    stamp = date_str.replace("-", "")   # YYYYMMDD — matches folder name prefix
    for folder in SOCIAL_DIR.iterdir():
        if not folder.is_dir() or not folder.name.startswith(stamp):
            continue
        imgs = [p for p in folder.iterdir()
                if p.suffix.lower() in IMG_EXTS and not p.name.startswith(".")]
        if not imgs:
            return None
        # Prefer the largest image (the promo, not an icon/thumbnail)
        return max(imgs, key=lambda p: p.stat().st_size)
    return None


def write_promo_files(shows):
    """Publish each public gig's promo to docs/shows/promos/<date>.<ext>."""
    PROMO_DIR.mkdir(exist_ok=True)
    count = 0
    for s in shows:
        if s.get("private"):
            continue
        src = _find_promo(s["date"])
        if not src:
            continue
        dest = PROMO_DIR / (s["date"] + src.suffix.lower())
        dest.write_bytes(src.read_bytes())
        count += 1
    return count


def main():
    if not XLSX_PATH.exists():
        sys.exit(f"Cannot find {XLSX_PATH}\nMake sure you're running from the repo root.")

    shows = build_shows()
    JSON_PATH.write_text(json.dumps(shows, indent=2, ensure_ascii=False))
    print(f"Wrote {len(shows)} shows to {JSON_PATH.name}")
    n = write_ics_files(shows)
    print(f"Wrote {n} calendar files to {ICS_DIR.relative_to(REPO_ROOT)}/")
    p = write_promo_files(shows)
    print(f"Published {p} promo images to {PROMO_DIR.relative_to(REPO_ROOT)}/")


if __name__ == "__main__":
    main()
